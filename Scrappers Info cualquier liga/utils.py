# Imports
import csv
import time
import openpyxl
from bs4 import BeautifulSoup
from selenium.common.exceptions import TimeoutException

# Constantes
nom_estadisticas = ['Goles esperados (xG)','Posesión de balón', 'Remates', 'Remates a puerta', 'Remates fuera', 'Remates rechazados',
                    'Tiros libres', 'Córneres', 'Fueras de juego', 'Saques de banda', 'Paradas', 'Faltas',
                    'Tarjetas rojas', 'Tarjetas amarillas', 'Pases totales', 'Pases completados', 'Tackles',
                    'Ataques', 'Ataques peligrosos']
titulos_estadisticas = (
                    'Date',
                    'HomeTeam',
                    'AwayTeam',
                    'HG',  # Home Goals
                    'AG',# Away Goals
                    'HxG',
                    'AxG',  
                    'HP',  # Home Possession
                    'AP',  # Away Possession
                    'HTS',  # Home Total Shots
                    'ATS',  # Away Total Shots
                    'HSI',  # Home Shots On Goal
                    'ASI',  # Away Shots On Goal
                    'HSO',  # Home Shots Off Goal
                    'ASO',  # Away Shots Off Goal
                    'HBS',  # Home Blocked Shots
                    'ABS',  # Away Blocked Shots
                    'HFK',  # Home Free Kicks
                    'AFK',  # Away Free Kicks
                    'HC',  # Home Corner Kicks
                    'AC',  # Away Corner Kicks
                    'HOFF',  # Home Offsides
                    'AOFF',  # Away Offsides
                    'HTI',  # Home Throw Ins
                    'ATI',  # Away Throw Ins
                    'HGS',  # Home Goalkeeper Saves
                    'AGS',  # Away Goalkeeper Saves
                    'HF',  # Home Fouls
                    'AF',  # Away Fouls
                    'HRC',  # Home Red Cards
                    'ARC',  # Away Red Cards
                    'HYC',  # Home Yellow Cards
                    'AYC',  # Away Yellow Cards
                    'HTP',  # Home Total Passes
                    'ATP',  # Away Total Passes
                    'HPC',  # Home Completed Passes
                    'APC',  # Away Completed Passes
                    'HT',  # Home Tackles
                    'AT',  # Away Tackles
                    'HA',  # Home Attacks
                    'AA',  # Away Attacks
                    'HDA',  # Home Dangerous Attacks
                    'ADA',  # Away Dangerous Attacks
                    'Resultado'
                )


# Funciones
# Crea un nuevo archivo en excel
def create_file(output_file):
    wb_init = openpyxl.Workbook()
    wb_init.save(output_file)
    wb_init.close()


# Obtiene el html con la lista de partidos
def obtener_lista_partidos(browser, link):
    soup = ""
    try:
        browser.get(link)
        while True:
            time.sleep(1)
            html = browser.page_source
            soup = BeautifulSoup(html, features="html.parser")
            if str(soup).__contains__("Mostrar más partidos"):
                time.sleep(1)
                browser.execute_script("arguments[0].click();",
                                       browser.find_element("xpath",'//*[@id="live-table"]/div[1]/div/div/a'))
            else:
                break
    except TimeoutException:
        print("I give up...")
    return soup


# Elimina la hoja por defecto
def eliminar_hoja_por_defecto(output_file):
    wb = openpyxl.load_workbook(output_file)
    del wb['Sheet']
    wb.save(output_file)
    wb.close()


# Convierte el archivo excel en un archivo CSV
def convertir_excel_a_csv(output_file):
    new_workbook = openpyxl.load_workbook(output_file)
    first_worksheet = new_workbook.active
    output_csv_file = csv.writer(open(output_file.split(".")[0] + ".csv", 'w', encoding="utf-8"), delimiter=",")
    for eachrow in first_worksheet.rows:
        output_csv_file.writerow([cell.value for cell in eachrow])
